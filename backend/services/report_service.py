from io import BytesIO
from pathlib import Path
import sys

VENDOR_PATH = Path(__file__).resolve().parents[1] / "vendor"
if VENDOR_PATH.exists() and str(VENDOR_PATH) not in sys.path:
    sys.path.insert(0, str(VENDOR_PATH))
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from backend.services.student_service import get_student_by_roll_number


def get_student_or_none(roll_number: str) -> dict | None:
    return get_student_by_roll_number(roll_number)


def subject_chart_image(student: dict) -> BytesIO:
    subjects = student.get("subjects", [])
    names = [subject["subject_name"] for subject in subjects]
    marks = [subject["total_marks"] for subject in subjects]

    fig, ax = plt.subplots(figsize=(7, 3.3))
    ax.bar(names, marks, color="#246bfe")
    ax.set_title("Subject-wise Marks")
    ax.set_ylabel("Marks")
    ax.set_ylim(0, 100)
    ax.tick_params(axis="x", rotation=30, labelsize=8)
    fig.tight_layout()

    image = BytesIO()
    fig.savefig(image, format="png", dpi=160)
    plt.close(fig)
    image.seek(0)
    return image


def cgpa_chart_image(student: dict) -> BytesIO:
    cgpa_rows = student.get("cgpa_by_semester", [])
    semesters = [row["semester"] for row in cgpa_rows]
    cgpa_values = [row["cgpa"] for row in cgpa_rows]

    fig, ax = plt.subplots(figsize=(7, 3))
    ax.plot(semesters, cgpa_values, marker="o", color="#0f9f6e", linewidth=2)
    ax.set_title("CGPA Trend")
    ax.set_xlabel("Semester")
    ax.set_ylabel("CGPA")
    ax.set_ylim(0, 10)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    image = BytesIO()
    fig.savefig(image, format="png", dpi=160)
    plt.close(fig)
    image.seek(0)
    return image


def table_style(header_fill="#111827") -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_fill)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d9e2ec")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f8fb")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
    )


def generate_student_pdf(student: dict) -> BytesIO:
    buffer = BytesIO()
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []

    story.append(Paragraph("Student Academic Performance Report", styles["Title"]))
    story.append(Paragraph(f"{student['name']} ({student['roll_number']})", styles["Heading2"]))
    story.append(Spacer(1, 0.15 * inch))

    profile_rows = [
        ["Class", student["class"], "Department", student["department"]],
        ["Semester", student["semester"], "Gender", student["gender"]],
        ["Attendance", f"{student['attendance_percentage']:.1f}%", "Overall CGPA", f"{student['overall_cgpa']:.2f}"],
        ["Overall Percentage", f"{student['overall_percentage']:.1f}%", "Grade", student["grade"]],
    ]
    profile_table = Table(profile_rows, colWidths=[1.35 * inch, 1.7 * inch, 1.5 * inch, 1.5 * inch])
    profile_table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d9e2ec"))]))
    story.append(profile_table)
    story.append(Spacer(1, 0.2 * inch))

    subject_rows = [["Subject", "Internal", "External", "Total"]]
    subject_rows.extend(
        [
            [
                subject["subject_name"],
                f"{subject['internal_marks']:.1f}",
                f"{subject['external_marks']:.1f}",
                f"{subject['total_marks']:.1f}",
            ]
            for subject in student.get("subjects", [])
        ]
    )
    subject_table = Table(subject_rows, colWidths=[2.4 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch])
    subject_table.setStyle(table_style())
    story.append(Paragraph("Subject Marks", styles["Heading3"]))
    story.append(subject_table)
    story.append(Spacer(1, 0.18 * inch))

    attendance_rows = [["Month", "Attendance %"]]
    attendance_rows.extend([[row["month"], f"{row['percentage']:.1f}%"] for row in student.get("attendance", [])])
    attendance_table = Table(attendance_rows, colWidths=[2.5 * inch, 2.0 * inch])
    attendance_table.setStyle(table_style())
    story.append(Paragraph("Attendance", styles["Heading3"]))
    story.append(attendance_table)
    story.append(Spacer(1, 0.18 * inch))

    cgpa_rows = [["Semester", "CGPA"]]
    cgpa_rows.extend([[row["semester"], f"{row['cgpa']:.2f}"] for row in student.get("cgpa_by_semester", [])])
    cgpa_table = Table(cgpa_rows, colWidths=[2.5 * inch, 2.0 * inch])
    cgpa_table.setStyle(table_style())
    story.append(Paragraph("CGPA Trend", styles["Heading3"]))
    story.append(cgpa_table)
    story.append(Spacer(1, 0.18 * inch))

    story.append(Paragraph("Charts", styles["Heading3"]))
    story.append(Paragraph("Subject-wise Marks Chart", styles["Heading4"]))
    story.append(Image(subject_chart_image(student), width=6.6 * inch, height=3.1 * inch))
    story.append(Spacer(1, 0.1 * inch))
    story.append(Paragraph("CGPA Line Chart", styles["Heading4"]))
    story.append(Image(cgpa_chart_image(student), width=6.6 * inch, height=2.8 * inch))

    doc.build(story)
    buffer.seek(0)
    return buffer


def style_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 34)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def generate_student_excel(student: dict) -> BytesIO:
    workbook = Workbook()
    profile = workbook.active
    profile.title = "Profile"
    profile.append(["Field", "Value"])
    profile_rows = [
        ("Name", student["name"]),
        ("Roll Number", student["roll_number"]),
        ("Class", student["class"]),
        ("Department", student["department"]),
        ("Semester", student["semester"]),
        ("Gender", student["gender"]),
        ("Attendance %", student["attendance_percentage"]),
        ("Overall Percentage", student["overall_percentage"]),
        ("Overall CGPA", student["overall_cgpa"]),
        ("Grade", student["grade"]),
    ]
    for row in profile_rows:
        profile.append(row)

    subjects = workbook.create_sheet("Subjects")
    subjects.append(["Subject", "Internal Marks", "External Marks", "Total Marks"])
    for subject in student.get("subjects", []):
        subjects.append(
            [
                subject["subject_name"],
                subject["internal_marks"],
                subject["external_marks"],
                subject["total_marks"],
            ]
        )

    attendance = workbook.create_sheet("Attendance")
    attendance.append(["Month", "Attendance %"])
    for row in student.get("attendance", []):
        attendance.append([row["month"], row["percentage"]])

    cgpa = workbook.create_sheet("CGPA Trend")
    cgpa.append(["Semester", "CGPA"])
    for row in student.get("cgpa_by_semester", []):
        cgpa.append([row["semester"], row["cgpa"]])

    for sheet in workbook.worksheets:
        style_sheet(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
